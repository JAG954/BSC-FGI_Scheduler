"""Workbook export helpers extracted from scheduler and data-import notebooks."""

from __future__ import annotations

from pathlib import Path
import shutil
import tempfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from bsc_fgi_scheduler.constants import FGI_TEAMS
from bsc_fgi_scheduler.paths import OUTPUT_DIR, ensure_output_dirs

SCHEDULER_TRACE_SHEETS = [
    "ChickenTracks",
    "Labor Allocation",
    "Moves Per Day",
    "Daily AP Status",
    "Exit Summary",
    "Active AP Status",
    "KPI Summary",
    "Team KPIs",
    "BTG structure",
    "BTG systems",
    "BTG declam",
    "BTG test",
]


def _style_workbook(path: Path, date_headers: set[str] | None = None):
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    index_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="BFBFBF")
    date_headers = date_headers or {"Date"}

    for ws in wb.worksheets:
        ws.freeze_panes = "B2"
        ws.sheet_view.showGridLines = False

        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if ws.max_column >= 1:
            for cell in ws["A"]:
                cell.fill = index_fill
                cell.font = Font(bold=True)
                if ws.cell(row=1, column=1).value in date_headers or ws.title.startswith("BTG"):
                    cell.number_format = "yyyy-mm-dd"
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin_gray)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col_idx in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col_idx).value
            if header_value in date_headers:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = "yyyy-mm-dd"

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for cell in col_cells:
                value = cell.value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 30)

        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def export_scheduler_trace(fgi, trace, output_dir: str | Path | None = None, output_file: str | Path | None = None) -> Path:
    output_dir = ensure_output_dirs(output_dir or OUTPUT_DIR)
    output_file = output_dir / "scheduler_trace_output.xlsx" if output_file is None else Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    chickentracks_df, labor_df, moves_df, btg_dfs = trace.to_dataframes()
    daily_status_df = fgi.get_daily_status_df()
    delivery_df = fgi.get_delivery_summary_df()
    active_status_df = fgi.get_active_ap_status_df()
    kpi_summary_df = fgi.get_kpi_summary_df(trace=trace)
    team_kpi_df = fgi.get_team_kpi_df(trace=trace)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        tmp_output = Path(tmp_file.name)

    try:
        with pd.ExcelWriter(tmp_output, engine="openpyxl") as writer:
            chickentracks_df.to_excel(writer, sheet_name="ChickenTracks")
            labor_df.to_excel(writer, sheet_name="Labor Allocation")
            moves_df.to_excel(writer, sheet_name="Moves Per Day")
            daily_status_df.to_excel(writer, sheet_name="Daily AP Status", index=False)
            delivery_df.to_excel(writer, sheet_name="Exit Summary", index=False)
            active_status_df.to_excel(writer, sheet_name="Active AP Status", index=False)
            kpi_summary_df.to_excel(writer, sheet_name="KPI Summary", index=False)
            team_kpi_df.to_excel(writer, sheet_name="Team KPIs", index=False)
            for team in FGI_TEAMS:
                df = btg_dfs.get(team, pd.DataFrame())
                df.to_excel(writer, sheet_name=f"BTG {team}"[:31])

        _style_workbook(tmp_output, date_headers={"Date", "FA_RO_Date", "Planned_B1R_Date", "Actual_Exit_Date"})
        shutil.copy2(tmp_output, output_file)

        exported = pd.ExcelFile(output_file).sheet_names
        missing_sheets = [sheet for sheet in SCHEDULER_TRACE_SHEETS if sheet not in exported]
        if missing_sheets:
            raise RuntimeError(f"Missing required export sheets: {missing_sheets}")
        return output_file
    finally:
        if tmp_output.exists():
            tmp_output.unlink()


def export_live_state(ap_df, location_df, labor_df, move_times_df=None, paint_schedule_df=None, output_path: str | Path | None = None) -> Path:
    output_path = Path(output_path) if output_path is not None else OUTPUT_DIR / "FGI_liveState.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        ap_df.to_excel(writer, sheet_name="ap_data", index=False)
        location_df.to_excel(writer, sheet_name="location_data", index=False)
        labor_df.to_excel(writer, sheet_name="labor_data", index=False)
        if move_times_df is not None:
            move_times_df.to_excel(writer, sheet_name="move_times")
        if paint_schedule_df is not None:
            paint_schedule_df.to_excel(writer, sheet_name="paint_schedule", index=False)

    _style_workbook(output_path, date_headers={"FA RO", "Date Online", "Date"})
    return output_path


def export_summary(summary_df: pd.DataFrame, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False)
    return output_path
